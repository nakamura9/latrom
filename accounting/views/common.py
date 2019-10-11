# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import datetime
import decimal
import json
import os
import urllib

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import UserPassesTestMixin
from django.http import HttpResponseRedirect, JsonResponse
from django.urls import reverse_lazy
from django.views.generic import DetailView, ListView, TemplateView
from django.views.generic.edit import (CreateView, DeleteView, FormView,
                                       UpdateView)
from django_filters.views import FilterView
from rest_framework import viewsets, generics
from django.shortcuts import get_object_or_404
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from formtools.wizard.views import SessionWizardView

from common_data.utilities import ContextMixin, apply_style, ConfigWizardBase
from common_data.views import PaginationMixin
from invoicing.models import Customer
from accounting import filters, forms, models, serializers
from accounting.views.reports.balance_sheet import BalanceSheet
from accounting.views.dash_plotters import expense_plot, revenue_vs_expense_plot
from employees.forms import EmployeeForm
from employees.models import Employee
import pygal
#constants
import openpyxl
import csv


CREATE_TEMPLATE = os.path.join('common_data', 'create_template.html')

class Dashboard( TemplateView):
    template_name = os.path.join('accounting', 'dashboard.html')

    def get(self, *args, **kwargs):
        config = models.AccountingSettings.objects.first()
        if config is None:
            config = models.AccountingSettings.objects.create(is_configured = False)
        if config.is_configured:
            return super().get(*args, **kwargs)
        else:
            return HttpResponseRedirect(reverse_lazy('accounting:config-wizard'))

    

class AsyncDashboard(TemplateView):
    template_name = os.path.join('accounting', 'async_dashboard.html')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        bsheet = BalanceSheet.common_context({})
        context['assets'] = bsheet['total_assets']
        context['liabilities'] = bsheet['current_liabilities_total'] + \
            bsheet['long_term_liabilities_total']
        context['equity'] = bsheet['equity_total']
        expense_chart = expense_plot()
        if expense_chart:
            context['expense_graph'] = expense_chart.render(is_unicode=True)
        else:
            context['expense_graph'] = "No expenses were recorded over the last 30 days"

        context['revenue_graph'] = revenue_vs_expense_plot().render(is_unicode=True)    
        return context


#############################################################
#                 JournalEntry Views                         #
#############################################################

# update and delete removed for security, only adjustments can alter the state 
# of an entry 

class JournalEntryCreateView( ContextMixin, CreateView):
    '''This type of journal entry has only one credit and one debit'''
    template_name = CREATE_TEMPLATE
    model = models.JournalEntry
    form_class = forms.SimpleJournalEntryForm
    
    extra_context = {
        "title": "Create New Journal Entry",
        "description": "Register Financial Transactions with the accounting system."
        }

    def get_success_url(self):
        return '/accounting/entry-detail/{}'.format(self.object.pk) 

class JournalEntryIframeView(ListView):
    template_name = os.path.join('accounting', 'journal', 'entry_list.html')
    paginate_by = 20
    def get_queryset(self):
        return models.JournalEntry.objects.filter(
            journal= models.Journal.objects.get(pk=self.kwargs['pk'])
        ).order_by('date').reverse()


class ComplexEntryView( ContextMixin, CreateView):
    '''This type of journal entry can have any number of 
    credits and debits. The front end page uses react to dynamically 
    alter the content of page hence the provided data from react is 
    sent to the server as urlencoded json in a hidden field called items[]
    
    '''
    template_name = os.path.join('accounting', 'compound_transaction.html')
    form_class= forms.ComplexEntryForm

    def post(self, request, *args, **kwargs):
        resp = super().post(request, *args, **kwargs)
        if not self.object:
            return resp

        for item in request.POST.getlist('items[]'):
            item_data = json.loads(urllib.parse.unquote(item))
            amount = decimal.Decimal(item_data['amount'])
                #incase the name includes '-' character
            pk, _ = item_data['account'].split("-")[:2] 
            account = models.Account.objects.get(
                    pk=int(pk))
            #make sure
            if int(item_data['debit']) == 1:
                self.object.debit(amount, account)
            else:
                self.object.credit(amount, account)

        return resp

class JournalEntryDetailView( DetailView):
    template_name = os.path.join('accounting', 'transaction_detail.html')
    model = models.JournalEntry

#############################################################
#                 Account  Views                            #
#############################################################
class AccountViewSet(viewsets.ModelViewSet):
    queryset = models.Account.objects.all()
    serializer_class = serializers.AccountSerializer


class AccountTransferPage( ContextMixin, CreateView):
    template_name = os.path.join('common_data','crispy_create_template.html')
    success_url = reverse_lazy('accounting:dashboard')
    form_class = forms.SimpleJournalEntryForm
    extra_context = {
        'title': 'Transfer between Accounts',
        'description': 'Move money directly between accounts using a simplified interface. Journal Entries are created automatically'
    }

class AccountCreateView( ContextMixin, CreateView):
    template_name = os.path.join('common_data','crispy_create_template.html')
    model = models.Account
    form_class = forms.AccountForm
    extra_context = {
        "title": "Create New Account",
        'description': "Use accounts to manage income and expenses in an intuitive way. A default chart of expenses is already implemented."}


class AccountUpdateView( ContextMixin, UpdateView):
    template_name = CREATE_TEMPLATE
    model = models.Account
    form_class = forms.AccountUpdateForm
    extra_context = {"title": "Update Existing Account"}


class AccountDetailView( DetailView):
    template_name = os.path.join('accounting', 'account','detail.html')
    model = models.Account 
    

class AccountCreditIframeView(ListView):
    template_name = os.path.join('accounting', 'account', 'entry_list.html')
    paginate_by = 20
    def get_queryset(self):
        return models.Credit.objects.filter(
            account= models.Account.objects.get(pk=self.kwargs['pk']),
            entry__draft=False
        ).order_by('pk')


class AccountDebitIframeView(ListView):
    template_name = os.path.join('accounting', 'account', 'entry_list.html')
    paginate_by = 20
    def get_queryset(self):
        return models.Debit.objects.filter(
            account= models.Account.objects.get(pk=self.kwargs['pk']),
            entry__draft=False
        ).order_by('pk')


class AccountListView( PaginationMixin, FilterView,  
        ContextMixin):
    template_name = os.path.join('accounting', 'account','list.html')
    filterset_class = filters.AccountFilter
    paginate_by = 20
    queryset = models.Account.objects.all()
    extra_context = {
        "title": "Chart of Accounts",
        'new_link': reverse_lazy('accounting:create-account'),
        'action_list': [
            {
                'link': reverse_lazy('accounting:import-accounts-from-excel'),
                'label': 'Import from Excel',
                'icon': 'file-excel'
            },
            {
                'link': reverse_lazy('accounting:bulk-create-accounts'),
                'label': 'Create Multiple Accounts ',
                'icon': 'file-alt'
            }
        ]
                }
    #model=models.Account

#############################################################
#                        Misc Views                         #
#############################################################

class TaxViewset(viewsets.ModelViewSet):
    queryset = models.Tax.objects.all()
    serializer_class = serializers.TaxSerializer

class TaxUpdateView( ContextMixin, UpdateView):
    form_class = forms.TaxUpdateForm
    model= models.Tax
    template_name = os.path.join('common_data','create_template.html')
    success_url = reverse_lazy('employees:dashboard')
    extra_context = {
        'title': 'Edit Sales Tax'
    }

class TaxCreateView( ContextMixin, CreateView):
    form_class = forms.TaxForm
    template_name = os.path.join('common_data','create_template.html')
    success_url = reverse_lazy('employees:dashboard')
    extra_context = {
        'title': 'Add Sales Taxes',
        'description': 'Sales taxes are used in orders and invoices.',
        
    }


class TaxListView( ContextMixin, PaginationMixin, FilterView):
    filterset_class = filters.TaxFilter
    template_name = os.path.join('accounting','tax_list.html')
    paginate_by = 20
    model = models.Tax
    extra_context = {
        'title': 'Sales Tax List',
        'new_link': reverse_lazy('accounting:create-tax')
    }

class TaxDeleteView( DeleteView):
    template_name = os.path.join('common_data', 'delete_template.html')
    success_url = reverse_lazy('employees:dashboard')
    model = models.Tax

class DirectPaymentFormView( ContextMixin, FormView):
    '''Uses a simple form view as a wrapper for a transaction in the journals
    for transactions involving two accounts.
    '''
    form_class = forms.DirectPaymentForm
    template_name = os.path.join('common_data', 'crispy_create_template.html')
    
    extra_context = {'title': 'Create Direct Payment'}

    def get_success_url(self, *args, **kwargs):
        return reverse_lazy('accounting:entry-detail', kwargs={
            'pk': models.JournalEntry.objects.latest('pk').pk + 1})

            
    def get_initial(self):
        if self.kwargs.get('supplier', None):
            return {
                'paid_to': self.kwargs['supplier']
            }
            
        return {}

    def post(self, request):
        resp = super(DirectPaymentFormView, self).post(request)
        form = self.form_class(request.POST)
        if form.is_valid():
            notes_string = """
                This payment was made out to: %s.
                the payment method used: %s \n """ % \
                (form.cleaned_data['paid_to'],
                    form.cleaned_data['method'])
            journal = models.Journal.objects.get(
                pk=4)#purchases journal
            j = models.JournalEntry.objects.create(
                memo=notes_string + form.cleaned_data['notes'],
                date=form.cleaned_data['date'],
                journal = journal,
                created_by=request.user
            )
            j.simple_entry(
                form.cleaned_data['amount'],
                form.cleaned_data['paid_to'].account,
                form.cleaned_data['account_paid_from'],
            )
        return resp

class AccountConfigView( UpdateView):
    '''
    Tabbed Configuration view for accounts 
    '''
    form_class = forms.ConfigForm
    template_name = os.path.join('accounting', 'config.html')
    success_url = reverse_lazy('accounting:dashboard')
    model = models.AccountingSettings

class DirectPaymentList( ContextMixin, TemplateView):
    template_name = os.path.join('accounting', 'direct_payment_list.html')
    extra_context = {
        'entries': lambda : models.JournalEntry.objects.filter(
           journal = models.Journal.objects.get(pk=4)) 
    }

#############################################################
#                    Journal Views                         #
#############################################################

class JournalCreateView( ContextMixin, CreateView):
    template_name = CREATE_TEMPLATE
    model = models.Journal
    form_class = forms.JournalForm
    extra_context = {
        "title": "Create New Journal",
        "description": 'A virtual document used to record all financial transactions in a business.'}

class JournalDetailView( DetailView):
    template_name = os.path.join('accounting', 'journal', 'detail.html')
    model = models.Journal

class JournalListView( ContextMixin, PaginationMixin, FilterView):
    template_name = os.path.join('accounting', 'journal', 'list.html')
    filterset_class = filters.JournalFilter
    paginate_by = 20
    extra_context = {
        "title": "Accounting Journals",
        'new_link': reverse_lazy('accounting:create-journal'),
        'action_list': [
            {
                'link': reverse_lazy('accounting:import-entries-from-excel'),
                'label': 'Import Entries from Excel',
                'icon': 'file-excel'
            },
            {
                'link': reverse_lazy('accounting:create-multiple-entries'),
                'label': 'Create Multiple Journal Entries ',
                'icon': 'file-alt'
            }
        ]}

    def get_queryset(self):
        return models.Journal.objects.all().order_by('name')

    

#########################################################
#                  Assets and Expenses                  #
#########################################################

class ExpenseAPIView(viewsets.ModelViewSet):
    queryset = models.Expense.objects.all()
    serializer_class = serializers.ExpenseSerializer


class CustomerExpenseAPIView(generics.ListAPIView):
    serializer_class = serializers.ExpenseSerializer
    
    def get_queryset(self):
        customer = Customer.objects.get(pk=self.kwargs['customer'])
        return models.Expense.objects.filter(customer=customer)

class AssetCreateView(ContextMixin,  CreateView):
    form_class = forms.AssetForm
    template_name = os.path.join('common_data','crispy_create_template.html')
    extra_context = {
        'title': 'Register New Asset',
        'description': 'Used to formally record valuable property belonging to the organization'
    }

    def post(self, *args, **kwargs):
        resp = super().post(*args, **kwargs)
        if self.object:
            self.object.create_entry()

        return resp
        
class AssetUpdateView(ContextMixin,  UpdateView):
    form_class = forms.AssetForm
    template_name = os.path.join('common_data','crispy_create_template.html')
    extra_context = {
        'title': 'Update Asset Data'
    }
    model = models.Asset


class AssetListView(ContextMixin,  PaginationMixin, 
        FilterView):
    template_name = os.path.join('accounting', 'asset_list.html')
    model = models.Asset
    filterset_class = filters.AssetFilter
    extra_context = {
        'title': 'List of Assets',
        'new_link': reverse_lazy('accounting:asset-create')
    }


class AssetDetailView( DetailView):
    template_name = os.path.join('accounting', 'asset_detail.html')
    model = models.Asset

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        chart = pygal.Line()
        start = self.object.init_date.year
        period = list(range(start, start + self.object.depreciation_period + 1))
        chart.x_labels = map(str, period)
        chart.add(self.object.name, [self.object.initial_value - \
            (i * self.object.annual_depreciation) \
                for i in range(self.object.depreciation_period + 1)])

        chart.title='Depreciation Plot'
        context['graph'] = chart.render(is_unicode=True)
        return context


class ExpenseCreateView(ContextMixin,  CreateView):
    form_class = forms.ExpenseForm
    template_name = os.path.join('common_data', 'crispy_create_template.html')
    extra_context = {
        'title': 'Record Expense',
        'description': 'Record costs incurred in the process of running a business.'
    }

    def post(self, *args, **kwargs):
        resp =super().post(*args, **kwargs)
        if self.object:
            self.object.create_entry()
        return resp

class ExpenseListView(ContextMixin,  PaginationMixin, 
        FilterView):
    template_name = os.path.join('accounting', 'expense','list.html')
    model = models.Expense
    filterset_class = filters.ExpenseFilter
    extra_context = {
        'title': 'List of Expenses',
        'new_link': reverse_lazy('accounting:expense-create'),
        'action_list': [
            {
                'link': reverse_lazy('accounting:recurring-expense-list'),
                'label': 'Manage Recurring Expenses',
                'icon': 'list-ul'
            },
            {
                'link': reverse_lazy('accounting:import-expenses'),
                'label': 'Import from Excel',
                'icon': 'file-excel'
            },
            {
                'link': reverse_lazy('accounting:create-multiple-expenses'),
                'label': 'Create Multiple Expenses',
                'icon': 'file-alt'
            }
        ]
    }


class ExpenseDetailView( DetailView):
    template_name = os.path.join('accounting', 'expense','detail.html')
    model = models.Expense

class ExpenseDeleteView( DeleteView):
    template_name = os.path.join('common_data', 'delete_template.html')
    model = models.Expense
    success_url = "/accounting/expense-list"

class RecurringExpenseCreateView(ContextMixin,  
        CreateView):
    form_class = forms.RecurringExpenseForm
    template_name = os.path.join('common_data','crispy_create_template.html')
    extra_context = {
        'title': 'Record Recurring Expense',
        'description': 'Record costs that occur periodically'
    }

class RecurringExpenseUpdateView(ContextMixin,  
        UpdateView):
    form_class = forms.RecurringExpenseForm
    template_name = os.path.join('common_data','crispy_create_template.html')
    extra_context = {
        'title': 'Update Recurring Expense'
    }
    model = models.RecurringExpense

class RecurringExpenseListView(ContextMixin,  
        PaginationMixin, FilterView):
    template_name = os.path.join('accounting', 'expense',
        'recurring','list.html')
    model = models.RecurringExpense
    filterset_class = filters.RecurringExpenseFilter
    extra_context = {
        'title': 'List of Recurring Expenses',
        'new_link': reverse_lazy('accounting:recurring-expense-create')
    }


class RecurringExpenseDetailView( DetailView):
    template_name = os.path.join('accounting', 'expense', 'recurring',
        'detail.html')
    model = models.RecurringExpense

class RecurringExpenseDeleteView( DeleteView):
    template_name = os.path.join('common_data', 'delete_template.html')
    model = models.RecurringExpense
    success_url = "/accounting/recurring-expense/list"



####################################################
#                  Bookeeper                       #
####################################################

class BookkeeperCreateView( CreateView):
    form_class = forms.BookkeeperForm
    template_name = CREATE_TEMPLATE
    success_url = reverse_lazy('accounting:bookkeeper-list')
    extra_context = {
        'title': 'Create Bookkeeper',
        'description': 'Assign An existing employee the role of Bookkeeper to manage the accounting system.'
    }

class BookkeeperUpdateView( UpdateView):
    form_class = forms.BookkeeperForm
    template_name = CREATE_TEMPLATE
    queryset = models.Bookkeeper.objects.all()
    success_url = reverse_lazy('accounting:bookkeeper-list')
    extra_context = {
        'title': 'Update Bookkeeper features'
    }

class BookkeeperListView( PaginationMixin ,FilterView):
    queryset = models.Bookkeeper.objects.filter(active=True)
    paginate_by = 20
    template_name = os.path.join('accounting', 'bookkeeper_list.html')
    extra_context = {
        'title': 'List of Bookkeepers',
        'new_link': reverse_lazy('accounting:create-bookkeeper')
    }
    filterset_class = filters.BookkeeperFilter


class BookkeeperDetailView( DetailView):
    model = models.Bookkeeper
    template_name = os.path.join('accounting', 'bookkeeper_detail.html')
    
    
class BookkeeperDeleteView(ContextMixin,  DeleteView):
    template_name = os.path.join('common_data', 'delete_template.html')
    model = models.Bookkeeper
    extra_context = {
        'title': 'Delete Bookkeeper'
    }

########################################################
#                   Migration Views                    #
########################################################

class BillCreateView(ContextMixin, CreateView):
    template_name =os.path.join('accounting', 'bill','create.html')
    extra_context = {
        'title': 'Create Bill',
        'description': 'Record money owed vendors for goods or services'
    }
    form_class = forms.BillForm

    def form_valid(self, form):
        resp = super().form_valid(form)

        data = json.loads(urllib.parse.unquote(
            form.cleaned_data['data']))
        
        
        for line in data:
            cat_string = line['category']
            #invert keys
            category = {i[1]: i[0] \
                for i in models.EXPENSE_CHOICES}.get(cat_string)

            models.BillLine.objects.create(
                bill=self.object,
                expense=models.Expense.objects.create(
                    debit_account=self.object.vendor.account,
                    date=self.object.date,
                    description=line['description'],
                    amount=line['amount'],
                    category=category
                )
            )
        self.object.create_entry()
        return resp

class BillUpdateView(ContextMixin, UpdateView):
    template_name =os.path.join('accounting', 'bill','update.html')
    form_class = forms.BillForm
    model = models.Bill
    

class BillListView(ContextMixin, PaginationMixin, FilterView):
    extra_context = {
        'title': 'List of Bills',
        'new_link': reverse_lazy('accounting:create-bill')
    }
    queryset = models.Bill.objects.all()
    filterset_class = filters.BillFilter
    paginate_by=20
    template_name = os.path.join('accounting', 'bill', 'list.html')

class BillDetailView(DetailView):
    template_name = os.path.join('accounting', 'bill', 'detail.html')
    model = models.Bill

class BillPaymentView(ContextMixin, CreateView):
    form_class = forms.BillPaymentForm
    template_name = os.path.join('common_data', 'crispy_create_template.html')
    extra_context = {
        'title': 'Pay Bill'
    }

    def get_initial(self):
        return {
            'bill': self.kwargs['pk']
        }

    def form_valid(self, form):
        resp = super().form_valid(form)
        self.object.create_entry()
        return resp

########################################################
#                   Currency Views                     #
########################################################


class CurrencyConverterView( TemplateView):
    template_name = os.path.join('accounting', 'currency_converter.html')

class CurrencyCreateView( CreateView):
    template_name = CREATE_TEMPLATE
    model = models.Currency
    fields = "__all__"
    success_url = "accounting/currency-converter"

class CurrencyUpdateView( UpdateView):
    template_name = CREATE_TEMPLATE
    model = models.Currency
    fields = "__all__"
    success_url = "accounting/currency-converter"


class CurrencyConversionLineCreateView( 
        CreateView):
    template_name = CREATE_TEMPLATE
    model = models.CurrencyConversionLine
    fields = "__all__"
    success_url = "accounting/currency-converter"

class CurrencyConversionLineUpdateView( 
        UpdateView):
    template_name = CREATE_TEMPLATE
    model = models.CurrencyConversionLine
    fields = "__all__"
    success_url = "accounting/currency-converter"

class CurrencyAPIView(viewsets.ModelViewSet):
    queryset = models.Currency.objects.all()
    serializer_class = serializers.CurrencySerializer

class AccountingSettingsAPIView(viewsets.ModelViewSet):
    queryset = models.AccountingSettings.objects.all()
    serializer_class = serializers.AccountingSettingsSerializer

class CurrencyConversionLineAPIView(viewsets.ModelViewSet):
    queryset = models.CurrencyConversionLine.objects.all()
    serializer_class = serializers.CurrencyConversionLineSerializer

class CurrencyConversionTableAPIView(viewsets.ModelViewSet):
    queryset = models.CurrencyConversionTable.objects.all()
    serializer_class = serializers.CurrencyConversionTableSerializer


class ExchangeTableCreateView(CreateView):
    # no get only post
    form_class = forms.ExchangeTableForm
    success_url = "/accounting/currency-converter/"
    template_name = CREATE_TEMPLATE

def update_reference_currency(request, table=None, currency=None):
    table = models.CurrencyConversionTable.objects.get(pk=table)
    currency = models.Currency.objects.get(pk=currency)
    table.reference_currency = currency
    table.save()

    return JsonResponse({'status': 'ok'})


def exchange_rate(request, line=None):
    line = models.CurrencyConversionLine.objects.get(pk=line)
    line.exchange_rate  = request.POST['rate']
    line.save()
    return JsonResponse({'status': 'ok'})

def create_exchange_table_conversion_line(request):
    table = models.CurrencyConversionTable.objects.get(
        pk=request.POST['table_id']) 
    currency = models.Currency.objects.get(
        pk=request.POST['currency_id']
    )
    models.CurrencyConversionLine.objects.create(
        currency = currency,
        exchange_rate = request.POST['rate'],
        conversion_table = table
    )
    return JsonResponse({'status': 'ok'})

def verify_entry(request, pk=None):
    entry = get_object_or_404(models.JournalEntry, pk=pk)
    entry.verify()
    return HttpResponseRedirect('/accounting/entry-detail/{}'.format(pk))


def employee_condition(self):
    return Employee.objects.all().count() == 0

def bookkeeper_condition(self):
    return models.Bookkeeper.objects.all().count() == 0

class ConfigWizard(ConfigWizardBase):
    template_name = os.path.join('accounting', 'wizard.html')
    form_list = [
        forms.ConfigForm, 
        EmployeeForm,
        forms.BookkeeperForm, 
        forms.TaxForm
    ]

    condition_dict = {
        '1': employee_condition,
        '2': bookkeeper_condition
    }

    config_class = models.AccountingSettings
    success_url = reverse_lazy('accounting:dashboard')

########################################################
#                   Migration Views                    #
########################################################

class ImportAccountsView(ContextMixin, FormView):
    extra_context = {
        'title': 'Import Accounts from Excel'
    }
    '''
    Takes csv file or excel file.
    Record the corresponding columns to the following fields:
        name
        description
        balance
        type
        balance_sheet_category
        code
    state the starting row(inclusive)
    state the ending row(inclusive)'''
    form_class = forms.AccountImportForm
    template_name = os.path.join('common_data', 'crispy_create_template.html')
    success_url = reverse_lazy('accounting:account-list')

    def form_valid(self, form):
        resp = super().form_valid(form)
        file = form.cleaned_data['file']
        if file.name.endswith('.csv'):
            #process csv 
            pass
        else:

            cols = [
                form.cleaned_data['name'],
                form.cleaned_data['balance'],
                form.cleaned_data['description'],
                form.cleaned_data['code'],
                form.cleaned_data['type'],
                form.cleaned_data['balance_sheet_category'],
            ]
            wb = openpyxl.load_workbook(file.file)
            try:
                ws = wb[form.cleaned_data['sheet_name']]
            except:
                ws = wb.active
            for row in ws.iter_rows(min_row=form.cleaned_data['start_row'],
                    max_row = form.cleaned_data['end_row'], 
                    max_col=max(cols)):
                models.Account.objects.create(
                    name=row[form.cleaned_data['name'] - 1].value,
                    balance=row[form.cleaned_data['balance'] - 1].value,
                    description=row[form.cleaned_data['description'] -1].value,
                    id=row[form.cleaned_data['code'] -1].value,
                    type=row[form.cleaned_data['type'] -1].value,
                    balance_sheet_category=row[form.cleaned_data['balance_sheet_category']-1].value
                )
        return resp
class BulkAccountCreateView(FormView):
    template_name = os.path.join('accounting', 'account', 'bulk_create.html')
    form_class = forms.BulkAccountsForm
    success_url = reverse_lazy('accounting:account-list')

    def form_valid(self, form):
        resp = super().form_valid(form)
        data = json.loads(urllib.parse.unquote(form.cleaned_data['data']))
        for line in data:
            models.Account.objects.create(
                name=line['name'],
                description=line['description'],
                id=line['code'],
                type=line['type'],
                balance=line['balance'],
                balance_sheet_category=line['balance_sheet_category']
            )

        return resp

class ImportTransactionView(ContextMixin, FormView):
    form_class = forms.ImportJournalEntryForm
    template_name = os.path.join('common_data','crispy_create_template.html')
    success_url = reverse_lazy('accounting:journal-list')

    extra_context = {
        'title': 'Import Journal Entries View'
        }

    def form_valid(self, form):
        resp = super().form_valid(self)

        fields = {
            'acc': form.cleaned_data['account'] -1,
            'date': form.cleaned_data['date'] -1,
            'memo': form.cleaned_data['memo'] -1,
            'entry_id': form.cleaned_data['entry_id'] -1,
            'credit': form.cleaned_data['credit'] -1,
            'debit': form.cleaned_data['debit'] -1
        }

        print(fields['acc'])
        file = form.cleaned_data['file']
        if file.name.endswith('.csv'):
            #process csv 
            pass
        else:

            cols = fields.values()
            wb = openpyxl.load_workbook(file.file)
            try:
                ws = wb[form.cleaned_data['sheet_name']]
            except:
                ws = wb.active
            for row in ws.iter_rows(min_row=form.cleaned_data['start_row'],
                    max_row = form.cleaned_data['end_row'], 
                    max_col=max(cols)+1):
                entry = None
                qs = models.JournalEntry.objects.filter(
                    id=row[fields['entry_id']].value)
                if qs.exists():
                    entry = qs.first()
                else:
                    date = None
                    if isinstance(row[fields['date']].value, str):
                        date = datetime.datetime.strptime(
                            row[fields['date']].value,
                            '%Y-%m-%d')
                    else: 
                        date = row[fields['date']].value.strftime('%Y-%m-%d')

                    entry= models.JournalEntry.objects.create(
                        journal=models.Journal.objects.get(id=5),
                        memo=row[fields['memo']].value,
                        date=date,
                        id=row[fields['entry_id']].value,
                    )
                
                qs = models.Account.objects.filter(id=row[fields['acc']].value)
                if not qs.exists():
                    messages.info(self.request, 
                        f'Account with ID {row[fields["acc"]].value} does not exist, entry skipped')
                    continue
                acc = qs.first()
                print(entry)
                if row[fields['credit']].value and \
                        row[fields['credit']].value > 0:
                    
                    models.Credit.objects.create(
                        amount=row[fields['credit']].value,
                        account=acc,
                        entry=entry
                    )
                if row[fields['debit']].value and \
                        row[fields['debit']].value > 0:
                    models.Credit.objects.create(
                        amount=row[fields['debit']].value,
                        account=acc,
                        entry=entry
                    )
                

        

        return resp


class CreateMultipleEntriesView(FormView):
    template_name = os.path.join('accounting', 'journal', 
        'multiple_create.html')
    form_class = forms.MultipleEntriesForm
    success_url = reverse_lazy('accounting:journal-list')

    def form_valid(self, form):
        resp = super().form_valid(form)

        data = json.loads(urllib.parse.unquote(form.cleaned_data['data']))
        for line in data:
            entry = models.JournalEntry.objects.create(
                date=line['date'],
                journal=models.Journal.objects.get(pk=5),
                memo=line['memo'],
            )
            if float(line['credit']) > 0:
                models.Credit.objects.create(
                    entry=entry,
                    amount=line['credit'],
                    account=models.Account.objects.get(
                        pk=line['account'].split('-')[0])
                )
            
            if float(line['debit']) > 0:
                models.Debit.objects.create(
                    entry=entry,
                    amount=line['debit'],
                    account=models.Account.objects.get(
                        pk=line['account'].split('-')[0])
                )
        return resp


class ImportExpensesView(ContextMixin, FormView):
    form_class = forms.ImportExpensesForm
    template_name = os.path.join('common_data','crispy_create_template.html')
    success_url = reverse_lazy('accounting:expense-list')

    extra_context = {
        'title': 'Import Journal Entries View'
        }

    def form_valid(self, form):
        resp = super().form_valid(self)

        acc = form.cleaned_data['account_paid_from']
        fields = {
            'desc': form.cleaned_data['description'] -1,
            'date': form.cleaned_data['date'] -1,
            'category': form.cleaned_data['category'] -1,
            'amt': form.cleaned_data['amount'] -1,
        }

        file = form.cleaned_data['file']
        if file.name.endswith('.csv'):
            #process csv 
            pass
        else:

            cols = fields.values()
            wb = openpyxl.load_workbook(file.file)
            try:
                ws = wb[form.cleaned_data['sheet_name']]
            except:
                ws = wb.active
            for row in ws.iter_rows(min_row=form.cleaned_data['start_row'],
                    max_row = form.cleaned_data['end_row'], 
                    max_col=max(cols)+1):
                cat_string = row[fields['category']].value
                #invert keys
                category = {i[1]: i[0] \
                    for i in models.EXPENSE_CHOICES}.get(cat_string, 17)
                date = row[fields['date']].value
                if isinstance(date, str):
                    date = datetime.datetime.strptime(date, '%d/%m/%Y')

                exp = models.Expense.objects.create(
                    debit_account=acc,
                    date=date,
                    description=row[fields['desc']].value,
                    amount=row[fields['amt']].value,
                    category=category
                )
                exp.create_entry()

        return resp


class CreateMultipleExpensesView(FormView):
    template_name = os.path.join('accounting', 'expense', 
        'create_multiple.html')
    form_class = forms.MultipleExpensesForm
    success_url = reverse_lazy('accounting:expense-list')

    def form_valid(self, form):
        resp = super().form_valid(form)
        acc = form.cleaned_data['account_paid_from']

        data = json.loads(urllib.parse.unquote(form.cleaned_data['data']))
        for line in data:
            cat_string = line['category']
            #invert keys
            category = {i[1]: i[0] \
                for i in models.EXPENSE_CHOICES}.get(cat_string)

            exp = models.Expense.objects.create(
                    debit_account=acc,
                    date=datetime.datetime.strptime(
                        line['date'], '%Y-%m-%d'),
                    description=line['description'],
                    amount=line['amount'],
                    category=category
                )
            exp.create_entry()
        return resp